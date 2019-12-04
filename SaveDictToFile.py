import codecs
import copy
import os
import csv
import unittest
from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook


SEPARATOR = ','
NEWLINE = '\r\n'


class SaveDictToFile():
    @classmethod
    def init(cls, data, filename, fieldnames):
        if fieldnames is None: fieldnames = []
        if not isinstance(data, (dict, list)): raise ValueError('Wrong data')
        if not isinstance(filename, str): raise ValueError('Wrong filename')
        if not isinstance(fieldnames, list): raise ValueError('Wrong fieldnames')
        for key in fieldnames:
            if not isinstance(key, str): raise ValueError('Wrong fieldnames')
        data = copy.deepcopy(data)
        fieldnames = cls.generate_fieldnames(data, fieldnames)
        return data, fieldnames

    @classmethod
    def generate_fieldnames(cls, data, fieldnames):
        new_fieldnames = []
        for each in data.values() if isinstance(data, dict) else data:
            if not isinstance(each, dict): raise ValueError('Wrong data')
            for key, value in each.items():
                if value and key not in new_fieldnames: new_fieldnames.append(str(key))
                if isinstance(value, float): each[key] = str(value).replace('.', ',')
        additional_fields = [str(x) for x in new_fieldnames if x not in fieldnames]
        cleared_fields = [str(x) for x in fieldnames if x not in new_fieldnames]
        if cleared_fields: print('deleted columns: ' + ', '.join(cleared_fields))
        return [str(x) for x in fieldnames if x in new_fieldnames] + additional_fields

    @classmethod
    def save_to_xlsx(cls, data, filename='', fieldnames=None):
        data, fieldnames = cls.init(data, filename, fieldnames)
        wb = Workbook()
        ws = wb.active
        ws.append(['#'] + fieldnames)
        for i, each in enumerate(data.values() if isinstance(data, dict) else data):
            line = [i + 1]
            for key in fieldnames:
                line.append(str(each.get(key, '')))
            ws.append(line)
        wb.save(f'{datetime.strftime(datetime.now(), "%Y-%m-%d_%H-%M")}_{filename}.xlsx')

    @classmethod
    def save_to_csv(cls, data, filename='', fieldnames=None):
        data, fieldnames = cls.init(data, filename, fieldnames)
        with codecs.open(f'{datetime.strftime(datetime.now(), "%Y-%m-%d_%H-%M")}_{filename}.csv', 'w', encoding='utf-8') as file:
            file.write('"#",' + SEPARATOR.join([f'"{x}"' for x in fieldnames]) + NEWLINE)
            for i, each in enumerate(data.values() if isinstance(data, dict) else data):
                line = [i + 1]
                for key in fieldnames:
                    line.append(str(each.get(key, '')).replace('"', '""'))
                file.write(SEPARATOR.join([f'"{x}"' for x in line]) + NEWLINE)

    @classmethod
    def save_to_csv_old(cls, data, filename='', fieldnames=None):
        data, fieldnames = cls.init(data, filename, fieldnames)
        with codecs.open(f'{datetime.strftime(datetime.now(), "%Y-%m-%d_%H-%M")}_{filename}.csv', 'w', encoding='utf-8') as file:
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            for each in data.values() if isinstance(data, dict) else data:
                for key in list(each).copy():
                    if key not in fieldnames: del each[key]
                for key in fieldnames:
                    if key not in each: each[key] = ''
                writer.writerow(each)

    @classmethod
    def save_to_files(cls, data, filename='', fieldnames=None):
        cls.save_to_xlsx(data, filename, fieldnames)
        cls.save_to_csv(data, filename, fieldnames)


class SaveDictToFileTests(unittest.TestCase, SaveDictToFile):
    data = {'first': {'1': '1\r\n1', '2': 22.2}, 'second': {'2': 33, '3': '"4""4', '4': ''}}

    def test_save_to_xlsx(self):
        file_name = f'{datetime.strftime(datetime.now(), "%Y-%m-%d_%H-%M")}_.xlsx'
        result = [['#', '1', '2', '3'], ['1', '1\r\n1', '22,2', 'None'], ['2', 'None', '33', '"4""4']]
        xlsx = []
        SaveDictToFile.save_to_xlsx(self.data)
        sheet = load_workbook(file_name).active
        for row in range(1, sheet.max_row + 1):
            row_data = []
            for column in range(1, sheet.max_column + 1):
                row_data.append(str(sheet.cell(row=row, column=column).value))
            xlsx.append(row_data)
        self.assertEqual(result, xlsx)
        os.remove(file_name)

    def test_save_to_csv(self):
        file_name = f'{datetime.strftime(datetime.now(), "%Y-%m-%d_%H-%M")}_.csv'
        result = '"#","1","2","3"\r\n"1","1\r\n1","22,2",""\r\n"2","","33","""4""""4"\r\n'
        SaveDictToFile.save_to_csv(self.data)
        with codecs.open(file_name, 'r', encoding='utf-8') as file:
            csv = file.read()
        self.assertEqual(result, csv)
        os.remove(file_name)

    def test_save_to_csv_old(self):
        file_name = f'{datetime.strftime(datetime.now(), "%Y-%m-%d_%H-%M")}_old.csv'
        result = '1,2,3\r\n"1\r\n1","22,2",\r\n,33,"""4""""4"\r\n'
        SaveDictToFile.save_to_csv_old(self.data, 'old')
        with codecs.open(file_name, 'r', encoding='utf-8') as file:
            csv = file.read()
        self.assertEqual(result, csv)
        os.remove(file_name)


if __name__ == '__main__':
    unittest.main()
